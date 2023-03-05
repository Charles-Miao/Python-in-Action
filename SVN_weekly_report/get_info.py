import re
import os
import openpyxl
from pathlib import Path

delp = 'D:\python_script\SVN_weekly_report\SVN.xlsx'
if os.path.exists(delp):
	os.remove(delp)

versions = []
modifiers = []
times = []
descriptions = []
models = []

versionRegex = re.compile('Versions:(\s*)?(\w+\d+)')
modifierRegex = re.compile('Modifier:(\s*)?(\w+)')
timeRegex = re.compile('Modified time:(\s*)?(\d{4}\S\d{2}\S\d{2}\s\d{2}\S\d{2}\S\d{2})')
descriptionRegex = re.compile('(-){72}\\n\\n(.*?)?\\n(-){72}\\n(-){21}detail messages', re.DOTALL)
modelRegex = re.compile('Index:(\s*)?E:/Molly/(.*)?/(.*)?\\n(=){67}')

p = Path('D:\python_script\SVN_weekly_report\.')
filepath = list(p.glob('*.txt'))

for i in range(len(filepath)):
	with open(filepath[i], 'r', encoding='gb18030', errors='ignore') as f:
		content = f.read()
		version = versionRegex.search(content)
		modifier = modifierRegex.search(content)
		time = timeRegex.search(content)
		description = descriptionRegex.search(content)
		model = modelRegex.search(content)
		if version == None or modifier == None or time == None or description == None or model == None:
			print('File Content Error: '+ str(filepath[i]))
			continue
		else:
			version = version.group(2)
			modifier = modifier.group(2)
			time = time.group(2)
			description = description.group(2)
			model = model.group(2).split('/',1)[0]
			versions.append(version)
			modifiers.append(modifier)
			times.append(time)
			descriptions.append(description)
			models.append(model)

try:
	workbook = openpyxl.load_workbook('SVN.xlsx')
except Exception as e:
	workbook = openpyxl.Workbook()
sheet_name = workbook.active

my_title = [ 'Model','Version', 'Modifer', 'Time', 'Description']
sheet_name.append(my_title)

for i in range(5):
	if i == 1:
		for j,value in enumerate(versions):
			sheet_name.cell(row=j+2, column=i+1, value=value)
	if i == 2:
		for k,value in enumerate(modifiers):
			sheet_name.cell(row=k+2, column=i+1, value=value)
	if i == 3:
		for m,value in enumerate(times):
			sheet_name.cell(row=m+2, column=i+1, value=value)
	if i == 4:
		for n,value in enumerate(descriptions):
			sheet_name.cell(row=n+2, column=i+1, value=value)
	if i == 0:
		for p,value in enumerate(models):
			sheet_name.cell(row=p+2, column=i+1, value=value)
workbook.save('SVN.xlsx')
workbook.close()