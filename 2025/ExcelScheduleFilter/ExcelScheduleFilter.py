import sys
import tkinter as tk
from tkinterdnd2 import *
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from tabulate import tabulate
from datetime import datetime, timedelta
import pandas as pd
from tkcalendar import DateEntry  # 导入DateEntry

# 定义更精确的宽泛红色 RGB 范围
LOW_RED_R = 128
HIGH_RED_R = 255
LOW_RED_G = 0
HIGH_RED_G = 128
LOW_RED_B = 0
HIGH_RED_B = 128


def float_to_datetime(excel_float):
    """
    将Excel日期浮点数转换为datetime对象。

    Args:
        excel_float (float): Excel日期浮点数，表示从1899年12月30日起的天数。

    Returns:
        datetime: 转换后的datetime对象。

    """
    base_date = datetime(1899, 12, 30)
    return base_date + timedelta(days=excel_float)


def is_redish(color_obj):
    """
    判断颜色对象是否为红色系。

    Args:
        color_obj (Union[str, Color]): 颜色对象，支持字符串格式和颜色类对象。

    Returns:
        bool: 如果是红色系则返回True，否则返回False。

    Raises:
        None

    备注：
        支持处理带alpha通道的颜色格式。
        1. Alpha通道表示颜色的透明度，数值范围通常为00（完全透明）到FF（完全不透明）。例如：
        #FFAABBCC：前两位FF为Alpha通道，后六位AABBCC为RGB值。
        #AABBCC：不包含Alpha通道，仅RGB。
        2. 为何需要去除Alpha通道
        在代码的is_redish函数中，判断颜色是否为红色系时，透明度信息无关紧要。若保留Alpha通道可能导致误判（例如#FFAABBCC和#AABBCC实际RGB值相同，但直接比较会因Alpha差异导致结果不同）。
    """
    if not color_obj:
        return False
    try:
        hex_str = color_obj.rgb if hasattr(color_obj, 'rgb') else str(color_obj)
        hex_str = hex_str.replace("FF", "", 1) if hex_str.startswith("FF") else hex_str  # 去除Alpha通道
        hex_str = hex_str.replace("#", "").strip()

        if len(hex_str) == 6:
            r = int(hex_str[0:2], 16)
            g = int(hex_str[2:4], 16)
            b = int(hex_str[4:6], 16)
        elif len(hex_str) == 8:  # 处理AARRGGBB格式
            r = int(hex_str[2:4], 16)
            g = int(hex_str[4:6], 16)
            b = int(hex_str[6:8], 16)
        else:
            return False

        return (
            LOW_RED_R <= r <= HIGH_RED_R and
            LOW_RED_G <= g <= HIGH_RED_G and
            LOW_RED_B <= b <= HIGH_RED_B
        )
    except (ValueError, AttributeError):
        return False


def extract_sixth_row(file_path, start_date, end_date):
    """
    从Excel文件中提取第六行，并根据指定的日期范围过滤结果。

    Args:
        file_path (str): Excel文件的路径。
        start_date (datetime): 日期范围的开始日期。
        end_date (datetime): 日期范围的结束日期。

    Returns:
        tuple: 返回一个包含两个元素的元组。
            - filtered_row (list): 过滤后的第六行数据。
            - rows_with_target_content (list): 包含目标内容的行列表。

    Raises:
        FileNotFoundError: 如果指定的文件不存在，则引发此异常。
        KeyError: 如果工作表“生产排程”不存在，则引发此异常。
        Exception: 如果发生其他未知错误，则引发此异常。

    """
    try:
        workbook = load_workbook(file_path,data_only=True)
        sheet = workbook['生产排程']
        sixth_row = []
        for cell in sheet[7]:
            if isinstance(cell.value, datetime):
                sixth_row.append(cell.value.strftime('%Y-%m-%d'))
            elif isinstance(cell.value, (int, float)):
                try:
                    date_obj = float_to_datetime(cell.value)
                    sixth_row.append(date_obj.strftime('%Y-%m-%d'))
                except ValueError:
                    sixth_row.append(cell.value)
            else:
                sixth_row.append(cell.value)
        # print(sixth_row)
        filtered_row = []
        for item in sixth_row:
            if isinstance(item, str) and '-' in item:
                try:
                    date = datetime.strptime(item, '%Y-%m-%d')
                    if start_date <= date <= end_date:
                        filtered_row.append(item)
                except ValueError:
                    continue

        rows_with_target_content = []
        columns = ['G', 'H', 'I', 'K', 'L', 'M', 'N', 'O']
        target_values = ["2","3","10","16"]
        for col_index, date in enumerate(filtered_row, start=1):
            sixth_row_col_index = sixth_row.index(date) + 1
            for row in sheet.iter_rows(min_row=1, min_col=sixth_row_col_index, max_col=sixth_row_col_index):
                for cell in row:
                    is_cell_red = is_redish(cell.fill.start_color) if cell.fill.start_color else False
                    is_text_red = is_redish(cell.font.color) if cell.font.color else False
                    if str(cell.value) in target_values and (is_cell_red or is_text_red):
                        row_content = []
                        for col in columns:
                            cell_value = sheet[f'{col}{cell.row}'].value
                            if col in ['K'] and isinstance(cell_value, (int, float)):
                                try:
                                    cell_value = float_to_datetime(cell_value).strftime('%Y-%m-%d')
                                except ValueError:
                                    pass
                            row_content.append(cell_value)
                        row_content.append(cell.value)
                        row_content.append(date)
                        conditions = []
                        if is_cell_red:
                            conditions.append("单元格颜色红色")
                        if is_text_red:
                            conditions.append("文本颜色红色")
                        condition_str = ", ".join(conditions)
                        row_content.append(condition_str)
                        row_content.append(cell.row)
                        rows_with_target_content.append(row_content)

        return filtered_row, rows_with_target_content

    except FileNotFoundError:
        print(f"错误：未找到文件 {file_path}。")
    except KeyError:
        print("错误：工作表 '生产排程' 不存在。")
    except Exception as e:
        print(f"发生未知错误：{e}")
    return None, None


def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file_path:
        # 更新文本框内容
        entry.delete(0, tk.END)
        entry.insert(0, file_path)


def process_file():
    """
    处理文件函数。

    Args:
        无

    Returns:
        无

    Raises:
        无

    该函数首先通过图形界面获取文件路径、开始日期和结束日期。
    如果未选择文件，将弹出警告提示用户选择文件。
    如果未输入开始日期或结束日期，将弹出警告提示用户输入日期。
    然后，将输入的日期字符串转换为 datetime 对象，并检查日期格式是否正确。
    如果日期格式不正确，将弹出错误提示用户输入正确的日期格式（YYYY-MM-DD）。
    如果开始日期晚于结束日期，将弹出错误提示用户更正日期。
    接着，调用 extract_sixth_row 函数提取指定日期范围内的数据，并返回过滤后的日期和包含目标内容的行。
    如果提取到数据，将打印过滤后的日期和包含目标内容的行，并将数据保存为 Excel 文件。
    最后，弹出信息框提示用户筛选结果已保存。
    """
    file_path = entry.get()
    if not file_path:
        messagebox.showwarning("警告", "请先选择文件")
        return
    start_date_str = start_entry.get_date().strftime('%Y-%m-%d')  # 获取选择的日期并转换为字符串
    end_date_str = end_entry.get_date().strftime('%Y-%m-%d')  # 获取选择的日期并转换为字符串
    if not start_date_str or not end_date_str:
        messagebox.showwarning("警告", "请输入开始日期和结束日期")
        return
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    except ValueError:
        messagebox.showerror("错误", "日期格式错误，请使用 YYYY-MM-DD 格式。")
        return

    if start_date > end_date:
        messagebox.showerror("错误", "开始日期不能晚于结束日期。")
        return

    filtered_dates, rows_with_target_content = extract_sixth_row(file_path, start_date, end_date)
    if filtered_dates is not None and rows_with_target_content is not None:
        print(f"第6行中 {start_date_str} 到 {end_date_str} 之间（包含这两天）的日期为：")
        print(tabulate([filtered_dates], tablefmt="grid"))
        print(f"内容为2、3、10、16且颜色为宽泛红色或文本颜色为宽泛红色的单元格所在行的G、H、I、K、L、M、N、O列的内容，最后四列添加匹配值、排产日期、筛选条件、原表格行号：")
        for row in rows_with_target_content:
            print(tabulate([row], tablefmt="grid"))

        columns = ['Project', 'LUXPC Model', 'LUXPC PN', '出货时间', 'Order Qty', '包装方式', 'Remark1', 'Remark2', '匹配值', '排产日期', '筛选条件', '原表格行号']
        df = pd.DataFrame(rows_with_target_content, columns=columns)

        output_file = f"筛选结果_{start_date_str}_{end_date_str}.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = workbook.active
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

        print(f"筛选结果已保存到 {output_file}")
        messagebox.showinfo("保存成功", f"筛选结果已成功保存")


root = TkinterDnD.Tk()
root.title("生产计划数据筛选工具")

# 创建文本框
entry = tk.Entry(root, width=50)
entry.pack(pady=20)

# 选择文件按钮
select_button = tk.Button(root, text="选择文件", command=select_file)
select_button.pack(pady=5)


def drop_enter(event):
    event.widget.focus_force()
    return event.action


def drop_leave(event):
    return event.action


def drop(event):
    file_path = event.data.strip('{}')
    entry.delete(0, tk.END)  # 清空文本框内容
    entry.insert(0, file_path)  # 在文本框显示文件路径
    return event.action


# 将拖放绑定到文本框
entry.drop_target_register(DND_FILES)
entry.dnd_bind('<<DropEnter>>', drop_enter)
entry.dnd_bind('<<DropLeave>>', drop_leave)
entry.dnd_bind('<<Drop>>', drop)

# 添加开始日期输入框
start_label = tk.Label(root, text="开始日期：")
start_label.pack()
today = datetime.today()
start_entry = DateEntry(root, width=20, date_pattern='yyyy-mm-dd', year=today.year, month=today.month, day=today.day)
start_entry.pack()

# 添加结束日期输入框
end_label = tk.Label(root, text="结束日期：")
end_label.pack()
today = datetime.today()
end_entry = DateEntry(root, width=20, date_pattern='yyyy-mm-dd', year=today.year, month=today.month, day=today.day)
end_entry.pack()

# 处理文件按钮
process_button = tk.Button(root, text="处理文件", command=process_file)
process_button.pack(pady=20)

root.mainloop()
